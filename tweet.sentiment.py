import openpyxl
import pandas as pd
import snscrape.modules.twitter as sntwitter
import spacy
import streamlit as st

# Load the Italian language model
nlp = spacy.load("it_core_news_sm")

#@st.cache
def scrape_tweets(topics, usernames, since, until):
    tweets = []
    # Collect tweets from each user
    for username in usernames:
        for topic in topics: 
            for i,tweet in enumerate(sntwitter.TwitterSearchScraper(f" {topic} from:{username} since: {since} until: {until}").get_items()):
                tweets.append([tweet.url, 
                                     tweet.date, 
                                     tweet.user.displayname, 
                                     tweet.content,
                                     tweet.user.username, 
                                     tweet.user.followersCount, 
                                     tweet.retweetedTweet, 
                                     tweet.quotedTweet,
                                     tweet.inReplyToUser,
                                     tweet.inReplyToTweetId,
                                     tweet.mentionedUsers
                                       ])

    # Classify the sentiment of each tweet
    for tweet in tweets:
        # Use the nlp object to create a Doc object
        # containing the text of the tweet
        doc = nlp(tweet.content)

        # Classify the sentiment of the tweet
        sentiment = doc.cats["POSITIVE"] - doc.cats["NEGATIVE"]

        if sentiment > 0:
            sentiment = "positive"
        elif sentiment == 0:
            sentiment = "neutral"
        else:
            sentiment = "negative"

        # Add the sentiment to the tweet
        tweet["sentiment"] = sentiment

    # Return the collected tweets
    return tweets

def export_to_excel(tweets):
    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the tweets to the Excel worksheet
    for i, tweet in enumerate(tweets):
        worksheet.cell(row=i+1, column=1).value = tweet["username"]
        worksheet.cell(row=i+1, column=2).value = tweet["date"]
        worksheet.cell(row=i+1, column=3).value = tweet["text"]
        worksheet.cell(row=i+1, column=4).value = tweet["sentiment"]

    # Save the Excel workbook
    workbook.save("tweets.xlsx")

# Add a title and introductory text
# Add a welcoming message and instructions

def main():
	

	# Title
    st.title("# Twitter Sentiment Analyzer")
    st.subheader("This app allows you to collect tweets from multiple users and analyze the sentiment of each tweet. Follow the instructions below to get started.")
    
    message = st.text_area("Enter Text","Type Here....")
    topics = st.text_input("Enter the topics of the users you want to collect tweets from (comma-separated):")
    topics = topics.split(",")
    usernames = st.text_input("Enter the usernames of the users you want to collect tweets from (comma-separated):")
    usernames = usernames.split(",")
    since = st.text_input("Enter the start date for the collection (YYYY-MM-DD):")
    until = st.text_input("Enter the end date for the collection (YYYY-MM-DD):")
    tweets = scrape_tweets(topics, usernames, since, until)
    st.table(tweets)

    # Ask the user if they want to export the tweets to an Excel file
    if st.checkbox("Export to Excel"):
        # Export the collected tweets to an Excel file
        export_to_excel(tweets)
    
    # Ask the user if they want to display the Excel table
    if st.checkbox("Display Excel table"):
        # Read the Excel file using pandas
        df = pd.read_excel("tweets.xlsx")
    
        # Display the Excel table
        st.dataframe(df)
        
    
    # Ask the user if they want to display the sentiment distribution
    if st.checkbox("Display sentiment distribution"):
        # Count the number of tweets with each sentiment
        sentiments = [tweet["sentiment"] for tweet in tweets]
        counts = {
            "positive": sentiments.count("positive"),
            "neutral": sentiments.count("neutral"),
            "negative": sentiments.count("negative"),
        }
    
        # Ask the user to choose between a bar chart and line plot
        chart_type = st.radio("Choose a chart type:", ("Bar chart", "Line plot"))
    
        # Display the sentiment distribution using the chosen chart type
        if chart_type == "Bar chart":
            st.bar_chart(counts)
        else:
            st.line_chart(counts)

if __name__ == '__main__':
	main()



